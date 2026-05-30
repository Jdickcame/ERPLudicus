import datetime
from decimal import Decimal

import openpyxl
from core.mixins import BranchAccessMixin
from core.pagination import StandardResultsSetPagination
from django.db import transaction
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from rest_framework import filters, status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import (
    Category,
    InventoryAdjustment,
    InventoryAdjustmentDetail,
    Kardex,
    PhysicalInventory,
    PhysicalInventoryDetail,
    Product,
    ProductRecipe,
    Stock,
    Tag,
    Transfer,
)
from .serializers import (
    CategorySerializer,
    InventoryAdjustmentSerializer,
    KardexSerializer,
    PhysicalInventorySerializer,
    ProductRecipeSerializer,
    ProductSerializer,
    StockSerializer,
    TagSerializer,
    TransferSerializer,
)


# --- 1. CATEGORÍAS Y ETIQUETAS ---
class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsAuthenticated]


class TagViewSet(viewsets.ModelViewSet):
    queryset = Tag.objects.all()
    serializer_class = TagSerializer
    permission_classes = [IsAuthenticated]


# --- 2. PRODUCTOS Y RECETAS ---
class ProductViewSet(BranchAccessMixin, viewsets.ModelViewSet):
    def get_queryset(self):
        from .models import Product

        qs = Product.objects.all().prefetch_related("tags", "category", "stocks")

        is_active = self.request.query_params.get("is_active", "")
        if is_active == "true":
            qs = qs.filter(is_active=True)

        branch_id = self.request.query_params.get("branch_id")
        for_pos = self.request.query_params.get("for_pos")
        for_purchase = self.request.query_params.get("for_purchase")
        search = self.request.query_params.get("search")

        # 👇 NUEVO PARÁMETRO: Capturamos la caja 👇
        cash_register_id = self.request.query_params.get("cash_register_id")

        if search:
            qs = qs.filter(
                Q(name__icontains=search)
                | Q(sku__icontains=search)
                | Q(category__name__icontains=search)
                | Q(area__name__icontains=search)
            )

        # 🔥 LA MAGIA DEL CATÁLOGO DINÁMICO 🔥
        # Si la petición viene para el POS, filtramos estrictamente por la sede
        if for_pos == "true" and branch_id:
            qs = qs.filter(
                product_type__in=[
                    "STOCKED",
                    "FINISHED",
                    "INTERMEDIATE",
                    "SERVICE",
                    "CONSUMABLE",
                ]
            )
            qs = qs.filter(
                stocks__branch_id=branch_id, stocks__is_active=True
            ).distinct()

            # 👇 EL CANDADO POR CATEGORÍA DE CAJA 👇
            if cash_register_id:
                from cash.models import CashRegister

                try:
                    caja = CashRegister.objects.get(id=cash_register_id)
                    categorias_permitidas = caja.allowed_categories.all()

                    if categorias_permitidas.count() > 0:
                        qs = qs.filter(category__in=categorias_permitidas)
                except CashRegister.DoesNotExist:
                    pass

        # Si viene un branch_id (desde la pantalla de Catálogo web)
        elif branch_id:
            user = self.request.user
            if hasattr(user, "branch") and user.branch:
                qs = qs.filter(
                    stocks__branch_id=branch_id, stocks__is_active=True
                ).distinct()
            else:
                pass  # El Admin ve todo

        if for_purchase == "true":
            qs = qs.filter(is_purchasable=True)
            qs = qs.filter(product_type__in=["STOCKED", "CONSUMABLE", "INTERMEDIATE"])

        return qs

    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    search_fields = ["name", "sku", "category__name", "area__name"]
    ordering_fields = ["name", "sku", "price", "created_at"]

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        if request.query_params.get("for_pos") == "true":
            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def perform_update(self, serializer):
        serializer.save()

    def perform_destroy(self, instance):
        instance.delete()

    @action(detail=False, methods=["get"])
    def choices(self, request):
        from .models import Product

        def format_opts(choices):
            return [{"value": k, "label": v} for k, v in choices]

        return Response(
            {
                "product_types": format_opts(Product.PRODUCT_TYPES),
                "uom_choices": format_opts(Product.UOM_CHOICES),
            }
        )

    @action(detail=False, methods=["get"])
    def debug_products(self, request):
        from .models import Product

        products = Product.objects.all().values(
            "id", "name", "product_type", "is_active"
        )
        return Response(
            {
                "total": products.count(),
                "products": list(products[:50]),
            }
        )

    @action(detail=False, methods=["get"])
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        branch_id = request.query_params.get("branch_id")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Catálogo de Productos"

        headers = [
            "SKU / Código",
            "Nombre del Producto",
            "Categoría",
            "Tipo",
            "Ud. Medida",
            "Precio Base (S/)",
        ]

        if branch_id:
            headers.extend(["Precio en Sede (S/)", "Estado en Sede"])

        ws.append(headers)

        # 👇 EL TRUCO: Pasamos los datos por tu Serializer para obtener
        # exactamente la misma data limpia que usa React
        serializer = self.get_serializer(queryset, many=True)

        for p in serializer.data:
            precio_base = float(p.get("price") or 0.0)

            row = [
                p.get("sku") or "-",
                p.get("name") or "Sin Nombre",
                p.get("category_name") or "Sin Categoría",
                p.get("type_display") or p.get("product_type", "-"),
                p.get("uom_display") or "-",
                precio_base,
            ]

            # Lógica dinámica si estamos viendo una sede
            if branch_id:
                stock = p.get("stock")
                if stock:
                    # Buscamos is_enabled (o is_active como respaldo)
                    is_enabled = stock.get("is_enabled", stock.get("is_active", False))
                    estado_sede = "Habilitado" if is_enabled else "Oculto"

                    precio_sede_str = stock.get("selling_price")
                    precio_sede = (
                        float(precio_sede_str) if precio_sede_str else precio_base
                    )
                else:
                    estado_sede = "Inactivo"
                    precio_sede = precio_base

                row.extend([precio_sede, estado_sede])

            ws.append(row)

        # ========================================================
        # 🎨 DISEÑO, BORDES Y AUTO-AJUSTE
        # ========================================================
        header_fill = PatternFill(
            start_color="1E293B", end_color="1E293B", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )

        # Pintamos la cabecera
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Fijamos los anchos
        ws.column_dimensions["A"].width = 15  # SKU
        ws.column_dimensions["B"].width = 45  # Nombre
        ws.column_dimensions["C"].width = 20  # Categoría
        ws.column_dimensions["D"].width = 15  # Tipo
        ws.column_dimensions["E"].width = 12  # UOM
        ws.column_dimensions["F"].width = 18  # Precio Base

        if branch_id:
            ws.column_dimensions["G"].width = 18  # Precio Sede
            ws.column_dimensions["H"].width = 15  # Estado

        # Formato de moneda para columnas de precio (Columna F, y G si hay sede)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row[5].number_format = '"S/" #,##0.00'  # Columna F (Índice 5)
            if branch_id:
                row[6].number_format = '"S/" #,##0.00'  # Columna G (Índice 6)

        # Retornamos el Excel construido
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="Catalogo_Productos.xlsx"'
        )
        wb.save(response)

        return response


class ProductRecipeViewSet(viewsets.ModelViewSet):
    queryset = ProductRecipe.objects.all()
    serializer_class = ProductRecipeSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        finished_product_id = self.request.query_params.get("finished_product")
        if finished_product_id:
            qs = qs.filter(finished_product_id=finished_product_id)
        return qs


# --- 3. STOCK ACTUAL ---
class StockViewSet(BranchAccessMixin, viewsets.ModelViewSet):
    serializer_class = StockSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["product__name", "product__sku", "product__category__name"]
    ordering_fields = ["id", "product__name", "quantity"]
    ordering = ["id"]

    def get_queryset(self):
        branch_id = self.request.query_params.get("branch_id")
        queryset = Stock.objects.filter(product__is_active=True).order_by("id")
        queryset = queryset.select_related("product", "product__category", "branch")

        if branch_id:
            queryset = queryset.filter(branch_id=branch_id)
        return queryset

    def create(self, request, *args, **kwargs):
        product_id = request.data.get("product")
        branch_id = request.data.get("branch")

        if not product_id or not branch_id:
            return Response(
                {"error": "Los campos 'product' y 'branch' son obligatorios."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        initial_qty = request.data.get("quantity", 0)

        stock_obj, created = Stock.objects.get_or_create(
            product_id=product_id,
            branch_id=branch_id,
            defaults={"quantity": initial_qty},
        )

        if not stock_obj.product.is_active:
            stock_obj.product.is_active = True
            stock_obj.product.save()

        serializer = self.get_serializer(stock_obj)
        return Response(
            serializer.data,
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK,
        )


# --- 4. KARDEX (HISTORIAL) ---
# --- 4. KARDEX (HISTORIAL) ---
class KardexViewSet(BranchAccessMixin, viewsets.ModelViewSet):
    serializer_class = KardexSerializer
    permission_classes = [IsAuthenticated]
    http_method_names = ["get", "head", "options"]
    pagination_class = StandardResultsSetPagination

    def get_queryset(self):
        queryset = (
            Kardex.objects.all()
            .order_by("-date")
            .select_related("product", "branch", "user")
        )

        branch_id = self.request.query_params.get("branch_id")
        if branch_id:
            queryset = queryset.filter(branch_id=branch_id)

        product_id = self.request.query_params.get("product")
        if product_id:
            queryset = queryset.filter(product_id=product_id)

        # 👇 AQUÍ ESTÁ LA MAGIA QUE FALTABA: EL FILTRO POR TIPO 👇
        movement_type = self.request.query_params.get("type")
        if movement_type:
            # Separamos por comas en caso de que vengan varios (Ej: "OUT_MERMA,OUT_COURTESY")
            types_list = movement_type.split(",")
            queryset = queryset.filter(type__in=types_list)

        start_date = self.request.query_params.get("start_date")
        end_date = self.request.query_params.get("end_date")

        if start_date:
            queryset = queryset.filter(date__gte=f"{start_date} 00:00:00")
        if end_date:
            queryset = queryset.filter(date__lte=f"{end_date} 23:59:59")

        return queryset


# --- 5. MOTOR DE AJUSTES (MERMAS, SOBRANTES, INICIAL) ---
class InventoryAdjustmentViewSet(BranchAccessMixin, viewsets.ModelViewSet):
    queryset = InventoryAdjustment.objects.all().order_by("-created_at")
    serializer_class = InventoryAdjustmentSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination

    def get_queryset(self):
        queryset = (
            Kardex.objects.all()
            .order_by("-date")
            .select_related("product", "branch", "user")
        )

        branch_id = self.request.query_params.get("branch_id")
        product_id = self.request.query_params.get("product")
        movement_type = self.request.query_params.get("type")

        if branch_id:
            queryset = queryset.filter(branch_id=branch_id)
        if product_id:
            queryset = queryset.filter(product_id=product_id)

        if movement_type:
            # 👇 AHORA ACEPTA VARIOS TIPOS SEPARADOS POR COMA 👇
            types_list = movement_type.split(",")
            queryset = queryset.filter(type__in=types_list)

        start_date = self.request.query_params.get("start_date")
        end_date = self.request.query_params.get("end_date")
        if start_date:
            queryset = queryset.filter(date__date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__date__lte=end_date)

        return queryset

    def list(self, request, *args, **kwargs):
        try:
            return super().list(request, *args, **kwargs)
        except Exception as e:
            import traceback

            print("ERROR EN LISTA DE AJUSTES:")
            print(traceback.format_exc())
            return Response(
                {"error": f"Error al cargar historial de ajustes: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def create(self, request, *args, **kwargs):
        try:
            with transaction.atomic():
                branch_id = request.data.get("branch_id")
                adj_type = request.data.get("type")
                reason = request.data.get("reason", "Ajuste manual")
                details = request.data.get("details", [])

                if not details:
                    return Response(
                        {"error": "Debe incluir al menos un producto"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                adjustment = InventoryAdjustment.objects.create(
                    branch_id=branch_id,
                    type=adj_type,
                    reason=reason,
                    created_by=request.user,
                )

                is_entry = adj_type in [
                    "MERMA_RETURN",
                    "ADJUST_IN",
                    "INITIAL",
                    "PRODUCTION",
                ]

                kardex_type_map = {
                    "MERMA_OUT": "OUT_MERMA",
                    "MERMA_RETURN": "IN_RETURN",
                    "INTERNAL": "OUT_ADJUSTMENT",
                    "ADJUST_IN": "IN_ADJUSTMENT",
                    "ADJUST_OUT": "OUT_ADJUSTMENT",
                    "INITIAL": "IN_ADJUSTMENT",
                    "PRODUCTION": "IN_PRODUCTION",
                }
                k_type = kardex_type_map.get(adj_type, "OUT_ADJUSTMENT")

                for item in details:
                    product = get_object_or_404(Product, id=item.get("product_id"))

                    if not product.manage_stock:
                        continue

                    qty = Decimal(str(item.get("quantity", 0)))
                    if qty <= Decimal("0"):
                        raise ValueError(
                            f"La cantidad de '{product.name}' debe ser mayor a 0."
                        )

                    stock, _ = Stock.objects.select_for_update().get_or_create(
                        branch_id=branch_id,
                        product=product,
                        defaults={
                            "quantity": Decimal("0"),
                            "average_cost": Decimal("0"),
                        },
                    )

                    stock_qty = (
                        Decimal(str(stock.quantity)) if stock.quantity else Decimal("0")
                    )
                    avg_cost = (
                        Decimal(str(stock.average_cost))
                        if stock.average_cost
                        else Decimal("0")
                    )

                    if not is_entry:
                        if stock_qty < qty:
                            raise ValueError(
                                f"Stock insuficiente para '{product.name}'. Intentas retirar {qty}, pero solo hay {stock_qty}."
                            )
                        costo_unitario = avg_cost
                        qty_kardex = -qty
                        new_qty = stock_qty - qty
                    else:
                        costo_unitario = Decimal(str(item.get("unit_cost", avg_cost)))
                        total_value_old = stock_qty * avg_cost
                        total_value_new = qty * costo_unitario
                        new_qty = stock_qty + qty

                        if new_qty > Decimal("0"):
                            stock.average_cost = (
                                total_value_old + total_value_new
                            ) / new_qty
                        else:
                            stock.average_cost = Decimal("0")

                        qty_kardex = qty

                    stock.quantity = new_qty
                    stock.save()

                    InventoryAdjustmentDetail.objects.create(
                        adjustment=adjustment,
                        product=product,
                        quantity=qty,
                        unit_cost=costo_unitario,
                    )

                    Kardex.objects.create(
                        branch_id=branch_id,
                        product=product,
                        type=k_type,
                        quantity=qty_kardex,
                        unit_cost=costo_unitario,
                        total_cost=qty * costo_unitario,
                        balance_quantity=stock.quantity,
                        balance_unit_cost=stock.average_cost,
                        balance_total_cost=stock.quantity * stock.average_cost,
                        user=request.user,
                        reference_document=f"AJUSTE #{adjustment.id}",
                        description=reason,
                    )

                    if adj_type == "PRODUCTION" and product.has_recipe:
                        ingredients = ProductRecipe.objects.filter(
                            finished_product=product
                        )

                        if not ingredients.exists():
                            raise ValueError(
                                f"El producto '{product.name}' no tiene una receta configurada para producirse."
                            )

                        for recipe_item in ingredients:
                            insumo = recipe_item.ingredient

                            recipe_qty = (
                                Decimal(str(recipe_item.quantity))
                                if recipe_item.quantity
                                else Decimal("0")
                            )
                            qty_to_deduct = qty * recipe_qty

                            stock_insumo, _ = (
                                Stock.objects.select_for_update().get_or_create(
                                    branch_id=branch_id,
                                    product=insumo,
                                    defaults={
                                        "quantity": Decimal("0"),
                                        "average_cost": Decimal("0"),
                                    },
                                )
                            )

                            costo_insumo = (
                                Decimal(str(stock_insumo.average_cost))
                                if stock_insumo.average_cost
                                else Decimal("0")
                            )
                            stock_ins_qty = (
                                Decimal(str(stock_insumo.quantity))
                                if stock_insumo.quantity
                                else Decimal("0")
                            )

                            stock_insumo.quantity = stock_ins_qty - qty_to_deduct
                            stock_insumo.save()

                            Kardex.objects.create(
                                branch_id=branch_id,
                                product=insumo,
                                type="OUT_PRODUCTION",
                                quantity=-qty_to_deduct,
                                unit_cost=costo_insumo,
                                total_cost=qty_to_deduct * costo_insumo,
                                balance_quantity=stock_insumo.quantity,
                                balance_unit_cost=stock_insumo.average_cost,
                                balance_total_cost=stock_insumo.quantity
                                * stock_insumo.average_cost,
                                user=request.user,
                                reference_document=f"PROD #{adjustment.id}",
                                description=f"Consumo automático para fabricar {qty}x {product.name}",
                            )

                serializer = self.get_serializer(adjustment)
                return Response(serializer.data, status=status.HTTP_201_CREATED)

        except ValueError as ve:
            return Response({"error": str(ve)}, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            import traceback

            print("ERROR CRÍTICO AL CREAR AJUSTE:")
            print(traceback.format_exc())
            return Response(
                {"error": f"Error interno en Ajustes: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


# --- 6. MOTOR DE TRASLADOS (ENTRE SEDES) ---
class TransferViewSet(BranchAccessMixin, viewsets.ModelViewSet):
    queryset = Transfer.objects.all().order_by("-created_at")
    serializer_class = TransferSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=["post"])
    @transaction.atomic
    def receive(self, request, pk=None):
        transfer = self.get_object()

        if transfer.status != "PENDING":
            return Response(
                {"error": "Solo se pueden recibir traslados pendientes"}, status=400
            )

        for detail in transfer.details.all():
            if not detail.product.manage_stock:
                continue

            stock_origin = Stock.objects.select_for_update().get(
                branch=transfer.origin_branch, product=detail.product
            )

            if stock_origin.quantity < detail.quantity:
                return Response(
                    {
                        "error": f"La sede origen ya no tiene stock suficiente de '{detail.product.name}' para completar este traslado (Stock actual: {stock_origin.quantity})."
                    },
                    status=400,
                )

            costo_traslado = stock_origin.average_cost
            stock_origin.quantity -= detail.quantity
            stock_origin.save()

            Kardex.objects.create(
                branch=transfer.origin_branch,
                product=detail.product,
                type="OUT_TRANSFER",
                quantity=-detail.quantity,
                unit_cost=costo_traslado,
                total_cost=detail.quantity * costo_traslado,
                balance_quantity=stock_origin.quantity,
                balance_unit_cost=stock_origin.average_cost,
                balance_total_cost=stock_origin.quantity * stock_origin.average_cost,
                user=request.user,
                reference_document=f"TRASLADO OUT #{transfer.id}",
            )

            stock_dest, _ = Stock.objects.select_for_update().get_or_create(
                branch=transfer.destination_branch,
                product=detail.product,
                defaults={"quantity": 0, "average_cost": costo_traslado},
            )

            total_old = stock_dest.quantity * stock_dest.average_cost
            total_new = detail.quantity * costo_traslado
            stock_dest.quantity += detail.quantity

            if stock_dest.quantity > 0:
                stock_dest.average_cost = (total_old + total_new) / stock_dest.quantity

            stock_dest.save()

            Kardex.objects.create(
                branch=transfer.destination_branch,
                product=detail.product,
                type="IN_TRANSFER",
                quantity=detail.quantity,
                unit_cost=costo_traslado,
                total_cost=detail.quantity * costo_traslado,
                balance_quantity=stock_dest.quantity,
                balance_unit_cost=stock_dest.average_cost,
                balance_total_cost=stock_dest.quantity * stock_dest.average_cost,
                user=request.user,
                reference_document=f"TRASLADO IN #{transfer.id}",
            )

        transfer.status = "COMPLETED"
        transfer.received_by = request.user
        transfer.save()

        return Response(
            {"message": "Traslado recibido exitosamente y stock actualizado"}
        )


class PhysicalInventoryViewSet(viewsets.ModelViewSet):
    queryset = PhysicalInventory.objects.all().order_by("-created_at")
    serializer_class = PhysicalInventorySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        branch_id = self.request.query_params.get("branch_id")
        if branch_id:
            qs = qs.filter(branch_id=branch_id)
        return qs

    def perform_create(self, serializer):
        start_date = serializer.validated_data.get("start_date")
        end_date = serializer.validated_data.get("end_date")
        branch = serializer.validated_data.get("branch")

        inventory = serializer.save(created_by=self.request.user)

        stocks_actuales = Stock.objects.filter(
            branch=branch, product__manage_stock=True, is_active=True
        ).select_related("product")

        detalles_a_crear = []

        if start_date and end_date:
            end_datetime = datetime.datetime.combine(end_date, datetime.time.max)
            if timezone.is_aware(timezone.now()):
                end_datetime = timezone.make_aware(end_datetime)

            for stock in stocks_actuales:
                ultimo_kardex_anterior = (
                    Kardex.objects.filter(
                        branch=branch, product=stock.product, date__date__lt=start_date
                    )
                    .order_by("-date", "-id")
                    .first()
                )

                stock_inicial = (
                    ultimo_kardex_anterior.balance_quantity
                    if ultimo_kardex_anterior
                    else Decimal("0.0")
                )

                movimientos_periodo = Kardex.objects.filter(
                    branch=branch,
                    product=stock.product,
                    date__date__gte=start_date,
                    date__lte=end_datetime,
                ).aggregate(
                    total_entradas=Sum("quantity", filter=Q(quantity__gt=0)),
                    total_salidas=Sum("quantity", filter=Q(quantity__lt=0)),
                )

                entradas = movimientos_periodo["total_entradas"] or Decimal("0.0")
                salidas = abs(movimientos_periodo["total_salidas"] or Decimal("0.0"))

                stock_teorico_final = stock_inicial + entradas - salidas

                ultimo_kardex_periodo = (
                    Kardex.objects.filter(
                        branch=branch,
                        product=stock.product,
                        date__date__gte=start_date,
                        date__lte=end_datetime,
                    )
                    .order_by("-date", "-id")
                    .first()
                )

                costo_cierre = (
                    ultimo_kardex_periodo.balance_unit_cost
                    if ultimo_kardex_periodo
                    else stock.average_cost
                )

                detalles_a_crear.append(
                    PhysicalInventoryDetail(
                        inventory=inventory,
                        product=stock.product,
                        initial_stock=stock_inicial,
                        total_inputs=entradas,
                        total_outputs=salidas,
                        system_stock=stock_teorico_final,
                        unit_cost=costo_cierre,
                        physical_stock=stock_teorico_final,
                    )
                )
        else:
            for stock in stocks_actuales:
                detalles_a_crear.append(
                    PhysicalInventoryDetail(
                        inventory=inventory,
                        product=stock.product,
                        initial_stock=stock.quantity,
                        total_inputs=0,
                        total_outputs=0,
                        system_stock=stock.quantity,
                        unit_cost=stock.average_cost,
                        physical_stock=stock.quantity,
                    )
                )

        PhysicalInventoryDetail.objects.bulk_create(detalles_a_crear)

    @action(detail=True, methods=["post"])
    def save_draft(self, request, pk=None):
        try:
            inventory = self.get_object()
            if inventory.status == "CLOSED":
                return Response(
                    {"error": "No se puede editar un inventario cerrado."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            detalles_modificados = request.data.get("details", [])

            for item in detalles_modificados:
                try:
                    detalle = PhysicalInventoryDetail.objects.get(
                        id=item.get("id"), inventory=inventory
                    )

                    val_fisico = item.get("physical_stock")
                    if val_fisico in [None, ""]:
                        val_fisico = detalle.system_stock

                    detalle.physical_stock = val_fisico

                    action_taken = item.get("action_taken")
                    if action_taken:
                        detalle.action_taken = action_taken

                    action_notes = item.get("action_notes")
                    if action_notes is not None:
                        detalle.action_notes = action_notes

                    detalle.save()
                except PhysicalInventoryDetail.DoesNotExist:
                    continue

            return Response({"message": "Borrador guardado correctamente."})

        except Exception as e:
            return Response(
                {"error": f"Error guardando borrador: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    @action(detail=True, methods=["post"])
    def close_inventory(self, request, pk=None):
        try:
            inventory = self.get_object()

            if inventory.status == "CLOSED":
                return Response(
                    {"error": "El inventario ya fue auditado y cerrado."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            draft_response = self.save_draft(request)
            if draft_response.status_code != 200:
                return draft_response

            with transaction.atomic():
                detalles = inventory.details.all()
                ref_str = getattr(inventory, "reference", f"ID-{inventory.id}")

                for detalle in detalles:
                    diff = (
                        Decimal(str(detalle.difference))
                        if detalle.difference
                        else Decimal("0")
                    )
                    unit_cost = (
                        Decimal(str(detalle.unit_cost))
                        if detalle.unit_cost
                        else Decimal("0")
                    )

                    if diff != Decimal("0") and detalle.action_taken == "ADJUST":
                        adj_type = "ADJUST_IN" if diff > 0 else "ADJUST_OUT"
                        cantidad_absoluta = abs(diff)

                        stock, _ = Stock.objects.select_for_update().get_or_create(
                            branch=inventory.branch,
                            product=detalle.product,
                            defaults={
                                "quantity": Decimal("0"),
                                "average_cost": unit_cost,
                            },
                        )

                        stock_qty = (
                            Decimal(str(stock.quantity))
                            if stock.quantity
                            else Decimal("0")
                        )
                        nueva_cantidad = stock_qty + diff

                        if nueva_cantidad < Decimal("0"):
                            raise ValueError(
                                f"No se puede ajustar '{detalle.product.name}'. "
                                f"Intentas restar {cantidad_absoluta}, pero el stock actual es solo {stock_qty}."
                            )

                        ajuste = InventoryAdjustment.objects.create(
                            branch=inventory.branch,
                            type=adj_type,
                            reason=f"Cuadre de Inv. {ref_str} - Ajuste de Sistema",
                            created_by=request.user,
                        )

                        InventoryAdjustmentDetail.objects.create(
                            adjustment=ajuste,
                            product=detalle.product,
                            quantity=cantidad_absoluta,
                            unit_cost=unit_cost,
                        )

                        stock.quantity = nueva_cantidad
                        stock.save()

                        tipo_kardex = "IN_ADJUSTMENT" if diff > 0 else "OUT_ADJUSTMENT"

                        avg_cost = (
                            Decimal(str(stock.average_cost))
                            if stock.average_cost
                            else Decimal("0")
                        )
                        final_qty = (
                            Decimal(str(stock.quantity))
                            if stock.quantity
                            else Decimal("0")
                        )

                        Kardex.objects.create(
                            branch=inventory.branch,
                            product=detalle.product,
                            date=timezone.now(),
                            type=tipo_kardex,
                            quantity=cantidad_absoluta
                            if diff > 0
                            else -cantidad_absoluta,
                            unit_cost=unit_cost,
                            total_cost=cantidad_absoluta * unit_cost,
                            balance_quantity=final_qty,
                            balance_unit_cost=avg_cost,
                            balance_total_cost=final_qty * avg_cost,
                            user=request.user,
                            description=f"Ajuste por Cuadre de Inventario {ref_str}",
                        )

            inventory.status = "CLOSED"
            inventory.closed_at = timezone.now()
            inventory.save()

            return Response(
                {
                    "message": "Inventario cerrado. Se aplicaron solo los ajustes seleccionados."
                }
            )

        except ValueError as ve:
            return Response({"error": str(ve)}, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            import traceback

            print(traceback.format_exc())
            return Response(
                {"error": f"Error Crítico en el Servidor (Python): {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
