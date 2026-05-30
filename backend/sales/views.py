import threading
from decimal import Decimal

import openpyxl
import requests
from branches.models import Branch

# Imports de Modelos
from cash.models import CashMovement, CashShift
from django.conf import settings
from django.contrib.auth import get_user_model
from django.db import transaction
from django.db.models import Q, Sum
from django.http import HttpResponse
from events.models import Event, EventRegistration
from inventory.models import Kardex, ProductRecipe, Stock
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# DRF Imports
from rest_framework import serializers, status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .invoice_service import InvoiceService
from .models import CreditNote, Customer, Sale
from .serializers import CreditNoteSerializer, CustomerSerializer, SaleSerializer

User = get_user_model()

# --- VIEWSETS ---


class CustomerViewSet(viewsets.ModelViewSet):
    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=["get"])
    def search_doc(self, request):
        doc_number = request.query_params.get("doc")
        if not doc_number:
            return Response({"error": "Falta el numero de documento"}, status=400)

        # 1. BÚSQUEDA LOCAL
        c = Customer.objects.filter(tax_id=doc_number).first()
        if c:
            return Response({"exists_local": True, "data": self.get_serializer(c).data})

        # 2. DETERMINAR EL TIPO DE DOCUMENTO
        doc_type = None
        if len(doc_number) == 8:
            doc_type = "DNI"
        elif len(doc_number) == 9:
            doc_type = "CEE"
        elif len(doc_number) == 11:
            doc_type = "RUC"

        if not doc_type:
            return Response(
                {"error": "Longitud invalida (DNI=8, CE=9, RUC=11)"}, status=400
            )

        customer_data = None

        # 3. INTENTO 1: APISPERU (Solo soporta DNI y RUC)
        if doc_type in ["DNI", "RUC"]:
            token_apisperu = getattr(settings, "APISPERU_CONSULTA_TOKEN", "")
            try:
                if doc_type == "DNI":
                    r = requests.get(
                        f"https://dniruc.apisperu.com/api/v1/dni/{doc_number}?token={token_apisperu}",
                        timeout=5,  # ⏱️ Tiempo optimizado
                    )
                    d = r.json()
                    if r.status_code == 200 and d.get("nombres"):
                        customer_data = {
                            "name": f"{d.get('nombres', '')} {d.get('apellidoPaterno', '')} {d.get('apellidoMaterno', '')}".strip(),
                            "document_type": "DNI",
                            "tax_id": doc_number,
                            "address": "PERU",
                        }
                elif doc_type == "RUC":
                    r = requests.get(
                        f"https://dniruc.apisperu.com/api/v1/ruc/{doc_number}?token={token_apisperu}",
                        timeout=5,
                    )
                    d = r.json()
                    if r.status_code == 200 and d.get("razonSocial"):
                        customer_data = {
                            "name": d.get("razonSocial", ""),
                            "document_type": "RUC",
                            "tax_id": doc_number,
                            "address": d.get("direccion", "PERU"),
                        }
            except Exception as e:
                print(f"ApisPeru fallo o demoro ({e}). Saltando a Factiliza...")

        # 4. INTENTO 2: FACTILIZA (DNI, RUC y CEE)
        if not customer_data:
            print("Consultando API de respaldo (Factiliza)...")
            token_factiliza = getattr(settings, "FACTILIZA_TOKEN", "")
            headers = {"Authorization": f"Bearer {token_factiliza}"}

            try:
                if doc_type == "DNI":
                    r = requests.get(
                        f"https://api.factiliza.com/v1/dni/info/{doc_number}",
                        headers=headers,
                        timeout=5,
                    )
                    resp = r.json()
                    if r.status_code == 200:
                        data = resp.get("data", resp)
                        nombres = data.get("nombres")

                        if nombres:
                            name_str = f"{data.get('nombres', '')} {data.get('apellido_paterno', '')} {data.get('apellido_materno', '')}".strip()
                            customer_data = {
                                "name": name_str,
                                "document_type": "DNI",
                                "tax_id": doc_number,
                                "address": data.get(
                                    "direccion_completa", data.get("direccion", "PERU")
                                ),
                            }

                elif doc_type == "RUC":
                    r = requests.get(
                        f"https://api.factiliza.com/v1/ruc/info/{doc_number}",
                        headers=headers,
                        timeout=5,
                    )
                    resp = r.json()
                    if r.status_code == 200:
                        data = resp.get("data", resp)
                        # Leemos la variable exacta según el JSON que proporcionaste
                        razon_social = data.get("nombre_o_razon_social") or data.get(
                            "razon_social"
                        )

                        if razon_social:
                            customer_data = {
                                "name": str(razon_social).strip(),
                                "document_type": "RUC",
                                "tax_id": doc_number,
                                "address": data.get(
                                    "direccion_completa", data.get("direccion", "PERU")
                                ),
                            }

                elif doc_type == "CEE":
                    r = requests.get(
                        f"https://api.factiliza.com/v1/cee/info/{doc_number}",
                        headers=headers,
                        timeout=5,
                    )
                    resp = r.json()
                    # CE no trae success=true, solo message=Exito
                    if r.status_code == 200 and resp.get("message") == "Exito":
                        data = resp.get("data", resp)
                        nombres = data.get("nombres")

                        if nombres:
                            name_str = f"{data.get('nombres', '')} {data.get('apellido_paterno', '')} {data.get('apellido_materno', '')}".strip()
                            customer_data = {
                                "name": name_str,
                                "document_type": "CE",
                                "tax_id": doc_number,
                                "address": "PERU",  # Extranjería no devuelve dirección exacta
                            }
            except Exception as e:
                print(f"Factiliza fallo o crasheo: {e}")

        # 5. RETORNAR RESULTADOS AL FRONTEND
        if customer_data and customer_data.get("name"):
            return Response({"exists_local": False, "data": customer_data}, status=200)

        return Response(
            {
                "error": "No se encontraron datos en SUNAT/RENIEC/MIGRACIONES. Por favor, ingrese los datos manualmente."
            },
            status=404,
        )


class SaleViewSet(viewsets.ModelViewSet):
    serializer_class = SaleSerializer
    permission_classes = [IsAuthenticated]
    queryset = Sale.objects.all().order_by("-id")

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user

        # 1. IDENTIFICAMOS EL ORIGEN DE LA PETICIÓN
        # Si no nos envían nada, asumimos que están en la caja registradora (pos)
        origin = self.request.query_params.get("origin", "pos")
        is_boss = user.is_superuser or getattr(user, "role", "") in ["ADMIN", "MANAGER"]

        # 2. LÓGICA MIXTA
        if origin == "web" and is_boss:
            # ==========================================
            # MODO WEB: El Admin está en su computadora viendo reportes
            # ==========================================
            branch_id = self.request.query_params.get("branch_id")
            if branch_id:
                queryset = queryset.filter(branch_id=branch_id)
            shift_id = self.request.query_params.get("shift_id")
            if shift_id:
                queryset = queryset.filter(shift_id=shift_id)

            # Capturamos y aplicamos los filtros
            search = self.request.query_params.get("search")
            start_date = self.request.query_params.get("start_date")
            end_date = self.request.query_params.get("end_date")
            doc_type = self.request.query_params.get("document_type")
            payment_method = self.request.query_params.get("payment_method")
            sunat_status = self.request.query_params.get("sunat_status")

            # 👇 NUEVO: Atrapamos el filtro de Caja 👇
            cash_register_id = self.request.query_params.get("cash_register_id")

            start_datetime = self.request.query_params.get("start_datetime")
            end_datetime = self.request.query_params.get("end_datetime")

            if search:
                # 🧠 Búsqueda Inteligente: Si el usuario pegó un documento con guion (Ej: F001-23 o FC11-100)
                if "-" in search:
                    partes = search.split("-")
                    serie_buscada = partes[0].strip()
                    numero_buscado = partes[1].strip()

                    # Filtramos exigiendo que coincidan tanto la serie como el número (de la venta O de la nota de crédito)
                    queryset = queryset.filter(
                        Q(
                            series__icontains=serie_buscada,
                            number__icontains=numero_buscado,
                        )
                        | Q(
                            credit_notes__series__icontains=serie_buscada,
                            credit_notes__number__icontains=numero_buscado,
                        )
                    ).distinct()
                else:
                    # Búsqueda normal por cliente o por partes sueltas (incluyendo Notas de Crédito)
                    queryset = queryset.filter(
                        Q(customer__name__icontains=search)
                        | Q(customer__tax_id__icontains=search)
                        | Q(series__icontains=search)
                        | Q(number__icontains=search)
                        | Q(credit_notes__series__icontains=search)
                        | Q(credit_notes__number__icontains=search)
                    ).distinct()

            if start_date:
                queryset = queryset.filter(date__date__gte=start_date)
            if end_date:
                queryset = queryset.filter(date__date__lte=end_date)
            if start_datetime:
                queryset = queryset.filter(date__gte=start_datetime)
            if end_datetime:
                queryset = queryset.filter(date__lte=end_datetime)

            if doc_type:
                if doc_type == "BOL":
                    queryset = queryset.filter(invoice_type_code="03")
                elif doc_type == "FAC":
                    queryset = queryset.filter(invoice_type_code="01")
                elif doc_type == "NTV":
                    queryset = queryset.filter(invoice_type_code="00")
                elif doc_type == "TICKET":
                    queryset = queryset.filter(invoice_type_code="99")
                elif doc_type == "NC":
                    queryset = queryset.filter(credit_notes__isnull=False).distinct()

            if payment_method:
                queryset = queryset.filter(
                    payments__payment_method=payment_method
                ).distinct()

            if sunat_status:
                if sunat_status == "PENDING":
                    # Si es pendiente, buscamos los PENDING, los nulos o los vacíos por seguridad
                    queryset = queryset.filter(
                        Q(sunat_status="PENDING")
                        | Q(sunat_status__isnull=True)
                        | Q(sunat_status="")
                    )
                else:
                    queryset = queryset.filter(sunat_status=sunat_status)

            # 👇 NUEVO: Aplicamos el filtro de Caja 👇
            if cash_register_id:
                queryset = queryset.filter(shift__cash_register_id=cash_register_id)

        else:
            # ==========================================
            # MODO POS: Cajero (o Admin operando la caja)
            # ==========================================
            current_shift = CashShift.objects.filter(user=user, status="OPEN").first()

            if current_shift:
                # Solo ve lo que ha cobrado en este turno
                queryset = queryset.filter(shift=current_shift)
            else:
                # Si no abrió caja, la tabla sale vacía por seguridad
                return queryset.none()

        return queryset

    def list(self, request, *args, **kwargs):
        # 1. Obtenemos las ventas ya filtradas por fecha, sede, etc.
        queryset = self.filter_queryset(self.get_queryset())

        # 2. 💰 LA MAGIA: Sumamos el campo 'total' de todas las ventas filtradas.
        # Filtramos por status="COMPLETED" para que no sume el dinero de las ventas Anuladas.
        total_sum = (
            queryset.filter(status="COMPLETED").aggregate(total_sum=Sum("total"))[
                "total_sum"
            ]
            or 0.0
        )

        # 3. Paginamos los resultados como siempre
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            response = self.get_paginated_response(serializer.data)
            # 4. 💉 Inyectamos el gran total en la respuesta JSON
            response.data["total_amount"] = total_sum
            return response

        # Por si acaso no hay paginación (fallback)
        serializer = self.get_serializer(queryset, many=True)
        return Response({"results": serializer.data, "total_amount": total_sum})

    def create(self, request, *args, **kwargs):
        # 1. Ejecuta la creación normal
        response = super().create(request, *args, **kwargs)

        # 2. Interceptamos la respuesta para buscar si se generó un ticket de evento
        sale_id = response.data.get("id")
        if sale_id:
            from events.models import EventRegistration

            ticket = EventRegistration.objects.filter(sale_id=sale_id).first()

            # 3. Si hay ticket, le mandamos los datos EXACTOS a React para el QR
            if ticket:
                response.data["generated_ticket_code"] = ticket.ticket_code
                response.data["generated_ticket_quantity"] = ticket.total_quantity

        return response

    def perform_create(self, serializer):
        # 1. EL TRUCO: Usamos initial_data en lugar de request.data para soportar envíos masivos
        data = serializer.initial_data

        # CAPTURAMOS EL ORIGEN PARA SEPARAR CAJA DE WEB 👇
        origin = self.request.query_params.get("origin", "pos")

        branch_id = data.get("branch_id")
        if not branch_id and hasattr(self.request.user, "branch"):
            branch_id = self.request.user.branch.id

        is_courtesy = str(data.get("is_courtesy", "false")).lower() == "true"
        supervisor = None
        supervisor_pin = data.get("supervisor_pin")

        if is_courtesy:
            if (
                getattr(self.request.user, "can_authorize_voids", False)
                or self.request.user.is_superuser
            ):
                supervisor = self.request.user
            else:
                # Si el PIN viene con prefijo "LOCAL:", significa que fue validado por el frontend (Dexie)
                if supervisor_pin and str(supervisor_pin).startswith("LOCAL:"):
                    parts = str(supervisor_pin).split(":")
                    if len(parts) >= 2:
                        supervisor_id = int(parts[1])
                        supervisor = User.objects.filter(
                            id=supervisor_id, is_active=True
                        ).first()
                elif supervisor_pin and supervisor_pin == "BYPASS":
                    pass  # Admin bypass - no supervisor needed
                else:
                    # Validación normal de PIN (online)
                    if not supervisor_pin:
                        raise serializers.ValidationError(
                            {"error": "No tienes permisos. Se requiere PIN."}
                        )

                    supervisor = User.objects.filter(
                        pin=supervisor_pin, can_authorize_voids=True
                    ).first()

                    if not supervisor:
                        raise serializers.ValidationError(
                            {"error": "PIN invalido o sin permisos."}
                        )

        with transaction.atomic():
            # 2. Seguimos usando la variable 'data'
            raw_invoice_type = data.get("invoice_type_code")
            is_nota_venta = raw_invoice_type == "00"
            is_factura = raw_invoice_type == "01"
            cid = data.get("customer")

            # 👇 SOLUCIÓN: Auto-crear cliente si viene de la Web sin ID 👇
            client_doc = data.get("customer_document")
            client_name = data.get("customer_name")
            doc_type_front = data.get("customer_type", "DNI")
            customer_obj = None

            if (
                (not cid or int(cid) < 0)
                and client_doc
                and client_doc not in ["", "00000000"]
            ):
                # Lo buscamos o lo creamos al instante en Postgres/MySQL
                customer_obj, _ = Customer.objects.get_or_create(
                    tax_id=client_doc,
                    defaults={
                        "name": client_name or "Cliente General",
                        "document_type": doc_type_front,
                        "address": "PERU",
                    },
                )
                # Le asignamos el nuevo ID REAL para el resto del proceso
                cid = customer_obj.id

            elif cid and int(cid) > 0:
                # Si el ID es positivo, es un cliente normal que ya existía
                customer_obj = Customer.objects.filter(id=cid).first()

            shift = None
            serie = ""
            tipo = ""
            # 🔥 SEPARAMOS LA LÓGICA WEB Y POS 🔥
            if origin == "pos":
                shift = (
                    CashShift.objects.filter(user=self.request.user, status="OPEN")
                    .order_by("-opened_at")
                    .first()
                )
                if not shift:
                    raise serializers.ValidationError(
                        {
                            "error": "Error de seguridad: El usuario no tiene ningun turno de caja registrado para emitir comprobantes."
                        }
                    )
                caja = shift.cash_register

                if is_courtesy:
                    serie = "T001"
                    tipo = "99"
                elif is_nota_venta:
                    serie = "NV01"
                    tipo = "00"
                else:
                    if cid:
                        if Customer.objects.filter(
                            pk=cid, document_type="RUC"
                        ).exists():
                            is_factura = True
                    serie = caja.factura_series if is_factura else caja.boleta_series
                    tipo = "01" if is_factura else "03"
            else:
                # SI ORIGEN ES WEB (NO HAY CAJA REGISTRADORA)
                if is_courtesy:
                    serie = "T001"
                    tipo = "99"
                elif is_nota_venta:
                    serie = "NV01"
                    tipo = "00"
                else:
                    if cid:
                        if Customer.objects.filter(
                            pk=cid, document_type="RUC"
                        ).exists():
                            is_factura = True

                    # LA WEB AHORA RESPETA SUS PROPIAS SERIES DESDE LA SEDE
                    branch_obj = Branch.objects.get(id=branch_id)

                    serie = (
                        branch_obj.web_factura_series
                        if is_factura
                        else branch_obj.web_boleta_series
                    )
                    tipo = "01" if is_factura else "03"

            # VALIDACIÓN CRÍTICA: Si es cortesía, el supervisor ES OBLIGATORIO
            if is_courtesy and supervisor is None:
                raise serializers.ValidationError(
                    {
                        "error": "ERROR DE SEGURIDAD: No se puede hacer una cortesia sin autorizacion valida."
                    }
                )

            # 4. Correlativo
            last = Sale.objects.filter(series=serie).order_by("-number").first()
            new_num = (int(last.number) + 1) if last else 1

            # 5. Guardar Venta
            sale = serializer.save(
                branch_id=branch_id,
                shift=shift,  # Si es 'web', shift guardará None
                is_courtesy=is_courtesy,
                authorized_by=supervisor,
                customer=customer_obj,
            )

            # RESPETAR EL CORRELATIVO DEL FRONTEND (ventas offline)
            offline_invoice_number = data.get("local_invoice_number")
            if offline_invoice_number and "-" in offline_invoice_number:
                # Extraer serie y correlativo del frontend
                parts = offline_invoice_number.split("-")
                if len(parts) == 2:
                    frontend_serie = parts[0]
                    frontend_number = parts[1].zfill(8)
                    # Solo respetar si la serie del frontend coincide con la calculada
                    if frontend_serie == serie:
                        existing = Sale.objects.filter(
                            series=serie, number=frontend_number
                        ).exists()
                        if not existing:
                            sale.series = frontend_serie
                            sale.number = frontend_number
                    # Para notas de venta, verificar si viene con serie NV
                    elif frontend_serie.startswith("NV") and is_nota_venta:
                        existing = Sale.objects.filter(
                            series=frontend_serie, number=frontend_number
                        ).exists()
                        if not existing:
                            serie = frontend_serie
                            sale.series = frontend_serie
                            sale.number = frontend_number

            offline_uuid = data.get("uuid")
            offline_date = data.get("date")

            if offline_uuid:
                sale.uuid = offline_uuid
                sale.is_synced = False  # Marca que vino con retraso

            # Si el POS se desconectó a las 2:00 PM y nos manda la venta a las 5:00 PM,
            # respetamos las 2:00 PM como hora real de venta.
            if offline_date:
                from django.utils.dateparse import parse_datetime
                from django.utils.timezone import is_aware, make_naive

                # Convertimos el string de React a un objeto datetime de Python
                dt = (
                    parse_datetime(offline_date)
                    if isinstance(offline_date, str)
                    else offline_date
                )

                # Si tiene zona horaria (es 'aware'), se la quitamos (lo hacemos 'naive')
                if dt and is_aware(dt):
                    dt = make_naive(dt)

                sale.date = dt
            # ------------------------------------------------

            # Solo sobreescribir correlativo si no vino del frontend
            if not (
                offline_invoice_number
                and "-" in offline_invoice_number
                and sale.series == serie
            ):
                sale.series = serie
                sale.number = str(new_num).zfill(8)
            sale.invoice_type_code = tipo
            sale.save()

            # 5. Detalles, Stock y Kardex (CON LÓGICA HÍBRIDA MTS/MTO/COMBOS)
            total_gravada, total_igv = 0, 0

            # DEFINIMOS LA FUNCIÓN RECURSIVA DENTRO DEL SCOPE
            def procesar_descuento_inventario(
                producto, cantidad_a_descontar, branch_id_local, costo_heredado=None
            ):

                # =======================================================
                # CASO 1: EL PRODUCTO MANEJA STOCK (MTS - Make To Stock)
                # (Aplica para Gaseosas, Insumos, y Productos Terminados Pre-Preparados)
                # =======================================================
                if producto.manage_stock:
                    st, _ = Stock.objects.get_or_create(
                        branch_id=branch_id_local,
                        product=producto,
                        defaults={"quantity": 0},
                    )
                    # if st.quantity < cantidad_a_descontar:
                    #     raise serializers.ValidationError(
                    #         {
                    #             "error": f"Stock insuficiente en el insumo/producto '{producto.name}'. "
                    #             f"Tienes {st.quantity} y necesitas {cantidad_a_descontar}."
                    #         }
                    #     )
                    costo_unitario = (
                        costo_heredado
                        if costo_heredado is not None
                        else st.average_cost
                    )
                    st.quantity -= cantidad_a_descontar
                    st.save()

                    Kardex.objects.create(
                        branch_id=branch_id_local,
                        product=producto,
                        date=sale.date,
                        type="OUT_SALE" if not is_courtesy else "OUT_COURTESY",
                        quantity=-cantidad_a_descontar,
                        unit_cost=costo_unitario,
                        total_cost=Decimal(str(cantidad_a_descontar))
                        * Decimal(str(costo_unitario)),
                        balance_quantity=st.quantity,
                        balance_unit_cost=st.average_cost,
                        balance_total_cost=Decimal(str(st.quantity))
                        * Decimal(str(st.average_cost)),
                        user=self.request.user,
                        description=f"Venta {sale.series}-{sale.number}"
                        if not is_courtesy
                        else f"Cortesia {sale.series}-{sale.number} (Aut: {supervisor.first_name if supervisor else 'Admin'})",
                    )
                    return costo_unitario

                # =======================================================
                # CASO 2: EL PRODUCTO NO MANEJA STOCK, PERO TIENE RECETA
                # (MTO - Make To Order, Cafés, Combos, Promociones)
                # =======================================================
                elif not producto.manage_stock and producto.has_recipe:
                    receta_items = ProductRecipe.objects.filter(
                        finished_product=producto
                    )
                    costo_acumulado_receta = Decimal("0.0")

                    for item in receta_items:
                        cant_ingrediente_total = item.quantity * cantidad_a_descontar
                        costo_ing = procesar_descuento_inventario(
                            item.ingredient, cant_ingrediente_total, branch_id_local
                        )
                        costo_acumulado_receta += Decimal(str(costo_ing)) * Decimal(
                            str(cant_ingrediente_total)
                        )

                    return (
                        costo_acumulado_receta / Decimal(str(cantidad_a_descontar))
                        if cantidad_a_descontar > 0
                        else Decimal("0.0")
                    )

                # =======================================================
                # CASO 3: SERVICIOS PUROS (Sin stock, sin receta)
                # =======================================================
                else:
                    return Decimal("0.0")  # Los servicios no tienen costo de inventario

            # --- PROCESAMOS CADA LÍNEA DE LA VENTA ---
            for d in sale.details.all():
                producto_vendido = d.product
                cantidad_vendida = Decimal(str(d.quantity))

                # Disparamos la función mágica
                costo_final_unitario = procesar_descuento_inventario(
                    producto_vendido, cantidad_vendida, branch_id
                )

                # Guardamos el costo real en el detalle de la venta para calcular utilidades luego
                d.unit_cost = costo_final_unitario

                # --- CALCULO DE IMPUESTOS POR LÍNEA ---
                if not is_courtesy:
                    sub = float(d.subtotal)
                    base = sub / 1.18
                    total_gravada += base
                    total_igv += sub - base

                d.save()

            sale.total_gravada = Decimal(str(total_gravada)) if not is_courtesy else Decimal("0")
            sale.total_igv = Decimal(str(total_igv)) if not is_courtesy else Decimal("0")

            # 🔥 CORRECCIÓN SIRE: Aplicar descuento a total_gravada y total_igv
            if sale.discount_amount and sale.discount_amount > 0:
                disc_base = sale.discount_amount / Decimal("1.18")
                disc_igv = sale.discount_amount - disc_base
                sale.total_gravada = round(sale.total_gravada - disc_base, 2)
                sale.total_igv = round(sale.total_igv - disc_igv, 2)

            if is_courtesy:
                sale.total = 0
                sale.status = "COMPLETED"
                sale.sunat_status = "ACCEPTED"

            # Si es Nota de Venta, la damos por aceptada internamente
            if is_nota_venta:
                sale.sunat_status = "ACCEPTED"
                sale.sunat_description = "Uso Interno"

            sale.save()

            event_id = data.get("event_id")
            event_id = data.get("event_id")
            if event_id:
                try:
                    # 🛡️ BLOQUEO DE FILA: select_for_update() impide que 2 compras simultáneas repitan el mismo código
                    evento = Event.objects.select_for_update().get(id=event_id)
                    detalles_recibidos = data.get("details", [])

                    # LÓGICA DE CENTRALIZACIÓN POR CATEGORÍA
                    total_entradas = 0
                    for d in detalles_recibidos:
                        from inventory.models import Product

                        prod_obj = Product.objects.filter(id=d.get("product")).first()

                        if (
                            prod_obj
                            and prod_obj.category
                            # RECONOCIMIENTO DE CATEGORIA
                            and prod_obj.category.name.upper()
                            in ["BOLETERÍA", "EVENTOS"]
                        ):
                            total_entradas += int(d.get("quantity", 0))

                    if total_entradas > 0:
                        prefix = f"E-{evento.branch.name[:3].upper()}"
                        last_ticket = (
                            EventRegistration.objects.filter(
                                ticket_code__startswith=prefix
                            )
                            .order_by("-id")
                            .first()
                        )

                        if last_ticket:
                            try:
                                last_num = int(
                                    last_ticket.ticket_code.replace(prefix, "")
                                )
                                new_code = f"{prefix}{str(last_num + 1).zfill(3)}"
                            except:  # noqa: E722
                                new_code = f"{prefix}001"
                        else:
                            new_code = f"{prefix}001"

                        # 👇 NUEVA MAGIA: AUTO-NUMERACIÓN DE PARTICIPANTES 👇
                        attendee_data = data.get("attendee_data")

                        if attendee_data and isinstance(attendee_data, list):
                            schema = evento.form_schema or []

                            for attendee in attendee_data:
                                perfil_aplicado = attendee.get("perfil_aplicado")

                                # Buscamos la configuración de ese perfil principal (Ej: "Menores")
                                matched_profile = None
                                for p in schema:
                                    if p.get("profileName") == perfil_aplicado:
                                        matched_profile = p
                                        break

                                if matched_profile:
                                    # Extraemos qué compró exactamente el usuario (Ej: "Kids")
                                    categoria_elegida = str(
                                        attendee.get("categoria_elegida", "")
                                    ).strip()

                                    # Extraemos los diccionarios de React
                                    category_codes = matched_profile.get(
                                        "categoryCodes", {}
                                    )
                                    current_codes_dict = matched_profile.get(
                                        "currentCodes", {}
                                    )

                                    # 1. Buscamos el inicio para esa categoría (Ej: 1000)
                                    start_code = int(
                                        category_codes.get(categoria_elegida) or 0
                                    )

                                    # 2. Vemos por dónde va el conteo de esa categoría
                                    current_code = int(
                                        current_codes_dict.get(categoria_elegida)
                                        or start_code
                                    )

                                    # 3. Incrementamos
                                    next_code = current_code + 1
                                    current_codes_dict[categoria_elegida] = next_code

                                    # Guardamos el diccionario de progreso actualizado
                                    matched_profile["currentCodes"] = current_codes_dict

                                    # Inyectamos el Código al formulario del Participante
                                    attendee["N° DORSAL"] = str(next_code)
                                else:
                                    attendee["N° DORSAL"] = f"{new_code}-GEN"

                            evento.form_schema = schema
                            evento.save(update_fields=["form_schema"])
                        # 👆 FIN DE LA MAGIA 👇

                        # Creamos el registro con el conteo real de "Boletería"
                        EventRegistration.objects.create(
                            event=evento,
                            sale=sale,
                            ticket_code=new_code,
                            schedule_selected=data.get("schedule_selected"),
                            operation_number=data.get("operation_number"),
                            observations=data.get("observations"),
                            total_quantity=total_entradas,
                            redeemed_quantity=0,
                            advisor=data.get("advisor"),
                            attendee_data=attendee_data,  # 👈 Ahora se guarda con los dorsales inyectados
                        )
                except Event.DoesNotExist:
                    pass

            # 6. REGISTRO EN CAJA CORREGIDO
            if origin == "pos" and shift and not is_courtesy:
                payments_data = data.get("payments", [])

                if not payments_data and hasattr(sale, "total"):
                    CashMovement.objects.create(
                        shift=shift,
                        user=self.request.user,
                        amount=sale.total,
                        movement_type="IN",
                        concept="SALE",
                        description=f"Venta {sale.series}-{sale.number}",
                        related_sale=sale,
                    )
                else:
                    for p_data in payments_data:
                        monto_pago = float(p_data.get("amount", 0))
                        metodo_pago = p_data.get("payment_method", "CASH")

                        if monto_pago > 0:
                            CashMovement.objects.create(
                                shift=shift,
                                user=self.request.user,
                                amount=monto_pago,
                                movement_type="IN",
                                concept="SALE",
                                description=f"Venta {sale.series}-{sale.number} ({metodo_pago})",
                                related_sale=sale,
                            )

            # 7. Facturación Electrónica a SUNAT
            # El candado final para que NV01 y T001 NO pasen
            if not is_courtesy and not is_nota_venta:
                sale.sunat_status = "PENDING"
                sale.save()

                def enviar_a_sunat(venta_id):
                    from django.db import connection

                    try:
                        venta = Sale.objects.get(id=venta_id)
                        InvoiceService(venta).generar_comprobante()
                    except Exception as e:
                        print(f"Error en hilo de SUNAT: {e}")
                    finally:
                        connection.close()

                if not getattr(self, "is_bulk_sync", False):
                    transaction.on_commit(
                        lambda: threading.Thread(
                            target=enviar_a_sunat, args=(sale.id,)
                        ).start()
                    )

    # NUEVA ACCIÓN: REINTENTO MANUAL DE ENVÍO A SUNAT
    @action(detail=True, methods=["post"])
    def send_sunat(self, request, pk=None):
        sale = self.get_object()

        if sale.invoice_type_code == "99":
            return Response(
                {
                    "error": "Los tickets internos no se envian a SUNAT.",
                    "success": False,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        if sale.sunat_status == "ACCEPTED":
            return Response(
                {"message": "El documento ya esta aceptado.", "success": True},
                status=status.HTTP_200_OK,
            )

        try:
            # Llamamos al servicio para que haga el trabajo sucio y guarde en BD
            InvoiceService(sale).generar_comprobante()

            # Recargamos el objeto desde la base de datos para ver qué estado le puso el InvoiceService
            sale.refresh_from_db()

            if sale.sunat_status == "ACCEPTED":
                return Response(
                    {
                        "success": True,
                        "message": "Documento enviado y aceptado por SUNAT correctamente.",
                    }
                )
            else:
                return Response(
                    {
                        "success": False,
                        "error": sale.sunat_description
                        or "El documento fue rechazado por SUNAT.",
                    }
                )

        except Exception as e:
            return Response(
                {"success": False, "error": f"Error interno al enviar: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    # EL MOTOR DE SINCRONIZACIÓN OFFLINE-FIRST
    @action(detail=False, methods=["post"])
    def bulk_sync(self, request):
        sales_data = request.data

        if not isinstance(sales_data, list):
            return Response(
                {
                    "error": "Formato invalido. Se esperaba una lista de ventas [{}, {}]."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        synced_count = 0
        errors = []
        successes = []  # 🌟 NUEVO: Lista para decirle a React exactamente cuáles ya están listas

        self.is_bulk_sync = True

        for index, sale_json in enumerate(sales_data):
            sale_uuid = sale_json.get("uuid")

            # 1. IDEMPOTENCIA: Si ya existe, le decimos al front que es un ÉXITO para que deje de enviarlo
            if sale_uuid and Sale.objects.filter(uuid=sale_uuid).exists():
                successes.append({"uuid": sale_uuid})
                continue

            serializer = self.get_serializer(data=sale_json)

            try:
                if serializer.is_valid():
                    self.perform_create(serializer)
                    synced_count += 1
                    successes.append({"uuid": sale_uuid})  # 🌟 Añadimos a éxitos
                else:
                    errors.append(
                        {"index": index, "uuid": sale_uuid, "errors": serializer.errors}
                    )
            except Exception as e:
                # 🛡️ EL SALVAVIDAS: Si a pesar de todo MySQL lanza el error 1062 (Duplicado), lo salvamos
                if "1062" in str(e) or "Duplicate entry" in str(e):
                    successes.append({"uuid": sale_uuid})
                else:
                    errors.append({"index": index, "uuid": sale_uuid, "errors": str(e)})

        return Response(
            {
                "message": "Sincronizacion finalizada",
                "synced_count": synced_count,
                "successes": successes,  # 🌟 NUEVO: Se lo enviamos a React
                "errors": errors,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=["get"])
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset()).prefetch_related(
            "payments", "customer", "credit_notes"
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historial de Ventas"

        # 1. CABECERAS
        headers = [
            "Fecha",  # A
            "Hora",  # B
            "Documento",  # C
            "Serie",  # D
            "Correlativo",  # E
            "RUC/DNI Cliente",  # F
            "Razón Social / Nombre",  # G
            "Subtotal",  # H
            "IGV",  # I
            "Total (S/)",  # J
            "Estado SUNAT",  # K
            "Referencia",  # L
            "Método de Pago",  # M
        ]
        ws.append(headers)

        for s in queryset:
            notas = s.credit_notes.all()
            es_anulada = len(notas) > 0

            # --- 1. PREPARAMOS LOS DATOS DE LA VENTA ORIGINAL ---
            if s.series.startswith("F"):
                tipo_doc = "Factura"
            elif s.series.startswith("B"):
                tipo_doc = "Boleta"
            elif s.series.startswith("NV") or s.series.startswith("N"):
                tipo_doc = "Nota de Venta"
            else:
                tipo_doc = "Ticket"

            pagos = s.payments.all()
            metodos_pago = (
                ", ".join([p.payment_method for p in pagos])
                if len(pagos) > 0
                else "EFECTIVO"
            )

            documento_cliente = s.customer.tax_id if s.customer else "-"
            nombre_cliente = s.customer.name if s.customer else "Público General"

            # El monto original SIEMPRE va en positivo para respetar el libro mayor
            total_final = float(s.total) if s.total else 0.0
            subtotal = total_final / 1.18
            igv = total_final - subtotal

            if tipo_doc in ["Factura", "Boleta"]:
                raw_status = s.sunat_status or "PENDIENTE"
                if raw_status == "ACCEPTED":
                    estado_label = "ACEPTADO"
                elif raw_status == "REJECTED":
                    estado_label = "RECHAZADO"
                else:
                    estado_label = "PENDIENTE"
            else:
                # Si es Ticket o Nota de Venta, no va a SUNAT
                estado_label = "NO APLICA"

            # --- 2. DIBUJAMOS LA VENTA ORIGINAL ---
            ws.append(
                [
                    s.date.strftime("%d/%m/%Y") if s.date else "-",
                    s.date.strftime("%H:%M") if s.date else "-",
                    tipo_doc,
                    s.series,
                    s.number,
                    documento_cliente,
                    nombre_cliente,
                    subtotal,
                    igv,
                    total_final,
                    estado_label,
                    "-",
                    metodos_pago,
                ]
            )

            # --- 3. DIBUJAMOS LA NOTA DE CRÉDITO (SI TIENE) EN NEGATIVO ---
            if es_anulada:
                for nc in notas:
                    # Las Notas de Crédito SÍ van a SUNAT, así que traducimos su estado
                    raw_nc_status = nc.sunat_status or "PENDIENTE"
                    if raw_nc_status == "ACCEPTED":
                        estado_nc = "ACEPTADO"
                    elif raw_nc_status == "REJECTED":
                        estado_nc = "RECHAZADO"
                    else:
                        estado_nc = "PENDIENTE"

                    ws.append(
                        [
                            nc.date.strftime("%d/%m/%Y") if nc.date else "-",
                            nc.date.strftime("%H:%M") if nc.date else "-",
                            "Nota de Credito",
                            nc.series,
                            nc.number,
                            documento_cliente,
                            nombre_cliente,
                            -subtotal,
                            -igv,
                            -total_final,
                            estado_nc,
                            f"Ref: {s.series}-{s.number}",
                            "DEVOLUCION",
                        ]
                    )

        # ========================================================
        # 🎨 DISEÑO, BORDES Y AUTO-AJUSTE
        # ========================================================
        header_fill = PatternFill(
            start_color="1E293B", end_color="1E293B", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )

        # Pintamos la cabecera
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Fijamos los anchos a mano
        anchos_columnas = {
            "A": 12,  # Fecha
            "B": 8,  # Hora
            "C": 15,  # Documento (más ancho por "Nota de Credito")
            "D": 8,  # Serie
            "E": 12,  # Correlativo
            "F": 15,  # RUC/DNI
            "G": 35,  # Razón Social
            "H": 12,  # Subtotal
            "I": 12,  # IGV
            "J": 15,  # Total
            "K": 15,  # Estado SUNAT
            "L": 18,  # Referencia
            "M": 15,  # Método de pago
        }
        for letra, ancho in anchos_columnas.items():
            ws.column_dimensions[letra].width = ancho

        # Aplicamos el formato de Moneda a las columnas H, I, J
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=8, max_col=10):
            for cell in row:
                cell.number_format = '"S/" #,##0.00'

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="Historial_Ventas_Ludicus.xlsx"'
        )
        wb.save(response)
        return response


class CreditNoteViewSet(viewsets.ModelViewSet):
    queryset = CreditNote.objects.all()
    serializer_class = CreditNoteSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        user = self.request.user

        # Creamos una variable booleana para saber si el usuario es "Jefe"
        is_boss = (
            user.is_superuser
            or user.role in ["ADMIN", "MANAGER"]
            or getattr(user, "can_authorize_voids", False)
        )

        # Si NO es jefe, le exigimos estrictamente un PIN
        if not is_boss:
            supervisor_pin = self.request.data.get("supervisor_pin")

            if not supervisor_pin:
                raise serializers.ValidationError(
                    {
                        "error": "No tienes permisos. Se requiere PIN de un Gerente/Admin para anular."
                    }
                )

            User = get_user_model()
            supervisor = User.objects.filter(
                pin=supervisor_pin,
                role__in=["ADMIN", "MANAGER"],
                # (Opcional) Si quieres que los que tienen can_authorize_voids=True también puedan dar su PIN, ponlo aquí con Q()
            ).first()

            if (
                not supervisor
                and not User.objects.filter(
                    pin=supervisor_pin, can_authorize_voids=True
                ).exists()
            ):
                raise serializers.ValidationError(
                    {
                        "error": "PIN invalido o el usuario no tiene permisos de autorizacion."
                    }
                )

        # 2. LÓGICA DE ANULACIÓN
        sale_id = self.request.data.get("sale")
        sale = Sale.objects.get(pk=sale_id)

        if sale.credit_notes.exists():
            raise serializers.ValidationError(
                {"error": "Esta venta ya fue anulada/modificada."}
            )

        with transaction.atomic():
            # 👇 LA SOLUCIÓN: SERIES SEPARADAS PARA NO QUEMAR CORRELATIVOS SUNAT 👇
            if sale.invoice_type_code == "01":
                serie_nc = "FC01"  # Nota de Crédito de Factura (Va a SUNAT)
            elif sale.invoice_type_code == "03":
                serie_nc = "BC01"  # Nota de Crédito de Boleta (Va a SUNAT)
            elif sale.invoice_type_code == "00":
                serie_nc = "NC01"  # Anulación interna de Nota de Venta
            else:
                serie_nc = "TC01"  # Anulación interna de Ticket (99)
            # 👆 FIN DE LA SOLUCIÓN 👆

            last = (
                CreditNote.objects.filter(series=serie_nc).order_by("-number").first()
            )
            new_num = str(int(last.number) + 1).zfill(8) if last else "00000001"

            note = serializer.save(
                sale=sale,
                series=serie_nc,
                number=new_num,
                note_type="07",
            )

            def procesar_devolucion_inventario(
                producto, cantidad_a_devolver, branch_local
            ):
                from inventory.models import ProductRecipe

                # CASO 1: MANEJA STOCK (MTS - Insumos o Productos Directos)
                if producto.manage_stock:
                    st, _ = Stock.objects.get_or_create(
                        branch=branch_local, product=producto, defaults={"quantity": 0}
                    )
                    st.quantity += cantidad_a_devolver
                    st.save()

                    Kardex.objects.create(
                        branch=branch_local,
                        product=producto,
                        date=note.date,
                        type="IN_RETURN",
                        quantity=cantidad_a_devolver,
                        unit_cost=st.average_cost,
                        total_cost=Decimal(str(cantidad_a_devolver))
                        * Decimal(str(st.average_cost)),
                        balance_quantity=st.quantity,
                        balance_unit_cost=st.average_cost,
                        balance_total_cost=Decimal(str(st.quantity))
                        * Decimal(str(st.average_cost)),
                        user=self.request.user,
                        description=f"Anulacion {sale.series}-{sale.number}",
                    )

                # CASO 2: TIENE RECETA (MTO - Hamburguesas, Combos)
                elif not producto.manage_stock and producto.has_recipe:
                    receta_items = ProductRecipe.objects.filter(
                        finished_product=producto
                    )
                    for item in receta_items:
                        cant_ingrediente_total = item.quantity * cantidad_a_devolver
                        # Llamada recursiva para devolver los ingredientes
                        procesar_devolucion_inventario(
                            item.ingredient, cant_ingrediente_total, branch_local
                        )

                # CASO 3: SERVICIOS (No se devuelve nada)
                else:
                    pass

            # 3. EJECUTAMOS LA LOGÍSTICA INVERSA POR CADA LÍNEA DE LA VENTA
            if note.note_type == "07":
                for detail in sale.details.all():
                    cantidad_devuelta = Decimal(str(detail.quantity))
                    procesar_devolucion_inventario(
                        detail.product, cantidad_devuelta, sale.branch
                    )

            # 4. CAJA (Reembolso de dinero)
            original_shift = sale.shift
            if original_shift and note.note_type == "07":
                CashMovement.objects.create(
                    shift=original_shift,
                    user=self.request.user,
                    amount=sale.total,
                    movement_type="OUT",
                    concept="REFUND",
                    description=f"Devolucion {sale.series}-{sale.number}",
                    related_sale=sale,
                )

            sale.status = "CANCELED"
            sale.save()

        if sale.invoice_type_code not in ["99", "00"]:
            # CREAMOS UN HILO (THREAD) PARA NO HACER ESPERAR AL FRONTEND
            def enviar_nota_async(nota_id):
                from django.db import connection

                try:
                    nota_bd = CreditNote.objects.get(id=nota_id)
                    InvoiceService(None).enviar_nota(nota_bd)
                except Exception as e:
                    print(f"Error en hilo de SUNAT (Nota de Credito): {e}")
                finally:
                    connection.close()

            # Disparamos el hilo en segundo plano al instante
            threading.Thread(target=enviar_nota_async, args=(note.id,)).start()
        else:
            # 👇 CORRECCIÓN FINAL PARA EL EXCEL 👇
            note.sunat_status = "NO APLICA"
            note.sunat_description = "Anulacion de Uso Interno"
            note.save()
